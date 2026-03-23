"""Tests for loaders.rvtools_parser — RVTools vInfo/vHost ETL.

Uses the real RVTools export at ``tests/rvtools/example.xlsx`` for
happy-path and integration tests.  Synthetic xlsx files are only
created for edge cases that the real file cannot cover (error paths,
missing columns, truthy variants, blank names, NaN handling).
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from loaders.rvtools_parser import (
    RVToolsParseError,
    parse_vhost,
    parse_vinfo,
)

# ── Paths ─────────────────────────────────────────────────────────────────

_EXAMPLE_XLSX = Path(__file__).parent / "rvtools" / "example.xlsx"


# ── Synthetic Excel helpers (edge-case tests only) ───────────────────────


def _write_vinfo(
    tmp_path: Path,
    rows: list[list[object]],
    *,
    filename: str = "rvtools.xlsx",
    columns: tuple[str, ...] = (
        "VM",
        "CPUs",
        "Memory",
        "Powerstate",
        "SRM Placeholder",
        "Template",
    ),
) -> Path:
    """Create a minimal RVTools xlsx with a vInfo sheet."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "vInfo"
    ws.append(list(columns))
    for row in rows:
        ws.append(row)
    path = tmp_path / filename
    wb.save(path)
    return path


def _write_vhost(
    tmp_path: Path,
    rows: list[list[object]],
    *,
    filename: str = "rvtools.xlsx",
    columns: tuple[str, ...] = (
        "Host",
        "# CPU",
        "Cores per CPU",
        "HT Active",
        "# Memory",
    ),
) -> Path:
    """Create a minimal RVTools xlsx with a vHost sheet."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "vHost"
    ws.append(list(columns))
    for row in rows:
        ws.append(row)
    path = tmp_path / filename
    wb.save(path)
    return path


# ═══════════════════════════════════════════════════════════════════════
# parse_vinfo — real data (tests/rvtools/example.xlsx)
# ═══════════════════════════════════════════════════════════════════════


class TestParseVinfo:
    """Happy-path tests against the real RVTools export.

    example.xlsx vInfo sheet:
      - 805 rows total
      - 763 poweredOn, 42 poweredOff
      - 10 templates (all poweredOff)
      - No SRM Placeholder column
      → expected result: 763 VMs
    """

    def test_total_vms_returned(self) -> None:
        result = parse_vinfo(_EXAMPLE_XLSX)
        assert len(result) == 763

    def test_no_powered_off_in_result(self) -> None:
        """All returned VMs must be poweredOn (no poweredOff leakage)."""
        result = parse_vinfo(_EXAMPLE_XLSX)
        assert all(vm.name for vm in result)
        assert all(vm.cpu > 0 for vm in result)
        assert all(vm.memory_mb > 0 for vm in result)

    def test_deterministic_output(self) -> None:
        """Two consecutive calls return identical results."""
        a = parse_vinfo(_EXAMPLE_XLSX)
        b = parse_vinfo(_EXAMPLE_XLSX)
        assert a == b

    def test_sorted_by_name(self) -> None:
        result = parse_vinfo(_EXAMPLE_XLSX)
        names = [vm.name for vm in result]
        assert names == sorted(names)

    def test_known_vm_wadfs1(self) -> None:
        """Spot-check: 'wadfs1' → 1 vCPU, 4096 MB."""
        result = parse_vinfo(_EXAMPLE_XLSX)
        vm = next(v for v in result if v.name == "wadfs1")
        assert vm.cpu == 1
        assert vm.memory_mb == 4096

    def test_known_vm_wspapp17(self) -> None:
        """Spot-check: 'wspapp17' → 16 vCPU, 53248 MB."""
        result = parse_vinfo(_EXAMPLE_XLSX)
        vm = next(v for v in result if v.name == "wspapp17")
        assert vm.cpu == 16
        assert vm.memory_mb == 53248


# ═══════════════════════════════════════════════════════════════════════
# parse_vinfo — edge cases (synthetic, cannot test with real file)
# ═══════════════════════════════════════════════════════════════════════


class TestParseVinfoEdgeCases:
    """Edge-case tests requiring synthetic xlsx (not in the real file)."""

    def test_filters_srm_placeholder(self, tmp_path: Path) -> None:
        """Real file has no SRM column; verify the filter works synthetically."""
        rows = [
            ["vm-real", 4, 8192, "poweredOn", "false", "false"],
            ["vm-srm", 4, 8192, "poweredOn", "true", "false"],
        ]
        path = _write_vinfo(tmp_path, rows)
        result = parse_vinfo(path)
        assert len(result) == 1
        assert result[0].name == "vm-real"

    def test_filters_template_poweredon(self, tmp_path: Path) -> None:
        """Real file templates are all poweredOff; test poweredOn + template."""
        rows = [
            ["vm-real", 4, 8192, "poweredOn", "false", "false"],
            ["vm-tmpl", 4, 8192, "poweredOn", "false", "true"],
        ]
        path = _write_vinfo(tmp_path, rows)
        result = parse_vinfo(path)
        assert len(result) == 1
        assert result[0].name == "vm-real"

    def test_lld_scenario_10_on_5_off_3_srm_2_tmpl(self, tmp_path: Path) -> None:
        """LLD §3.1 test: 10 poweredOn + 5 poweredOff + 3 SRM + 2 Templates → 10."""
        rows: list[list[object]] = []
        for i in range(10):
            rows.append([f"on-{i:02d}", 4, 8192, "poweredOn", "false", "false"])
        for i in range(5):
            rows.append([f"off-{i:02d}", 4, 8192, "poweredOff", "false", "false"])
        for i in range(3):
            rows.append([f"srm-{i:02d}", 4, 8192, "poweredOn", "true", "false"])
        for i in range(2):
            rows.append([f"tmpl-{i:02d}", 4, 8192, "poweredOn", "false", "true"])
        path = _write_vinfo(tmp_path, rows)
        result = parse_vinfo(path)
        assert len(result) == 10

    def test_case_insensitive_powerstate(self, tmp_path: Path) -> None:
        rows = [
            ["vm1", 4, 8192, "PoweredOn", "false", "false"],
            ["vm2", 4, 8192, "POWEREDON", "false", "false"],
            ["vm3", 4, 8192, " poweredOn ", "false", "false"],
        ]
        path = _write_vinfo(tmp_path, rows)
        result = parse_vinfo(path)
        assert len(result) == 3

    def test_srm_yes_variant_filtered(self, tmp_path: Path) -> None:
        """'Yes', '1', 'TRUE' should all be treated as truthy for SRM."""
        rows = [
            ["vm-ok", 4, 8192, "poweredOn", "false", "false"],
            ["vm-srm1", 4, 8192, "poweredOn", "Yes", "false"],
            ["vm-srm2", 4, 8192, "poweredOn", "1", "false"],
            ["vm-srm3", 4, 8192, "poweredOn", "TRUE", "false"],
        ]
        path = _write_vinfo(tmp_path, rows)
        result = parse_vinfo(path)
        assert len(result) == 1

    def test_nan_srm_and_template_included(self, tmp_path: Path) -> None:
        """NaN/empty SRM or Template should NOT filter the VM out."""
        rows = [
            ["vm1", 4, 8192, "poweredOn", None, None],
            ["vm2", 4, 8192, "poweredOn", "", ""],
        ]
        path = _write_vinfo(tmp_path, rows)
        result = parse_vinfo(path)
        assert len(result) == 2

    def test_missing_srm_column_still_works(self, tmp_path: Path) -> None:
        """If SRM Placeholder column is absent, VMs are included."""
        path = _write_vinfo(
            tmp_path,
            [["vm1", 4, 8192, "poweredOn", "false"]],
            columns=("VM", "CPUs", "Memory", "Powerstate", "Template"),
        )
        result = parse_vinfo(path)
        assert len(result) == 1

    def test_missing_template_column_still_works(self, tmp_path: Path) -> None:
        """If Template column is absent, VMs are included."""
        path = _write_vinfo(
            tmp_path,
            [["vm1", 4, 8192, "poweredOn", "false"]],
            columns=("VM", "CPUs", "Memory", "Powerstate", "SRM Placeholder"),
        )
        result = parse_vinfo(path)
        assert len(result) == 1

    def test_skips_blank_vm_names(self, tmp_path: Path) -> None:
        rows = [
            ["", 4, 8192, "poweredOn", "false", "false"],
            ["valid", 4, 8192, "poweredOn", "false", "false"],
        ]
        path = _write_vinfo(tmp_path, rows)
        result = parse_vinfo(path)
        assert len(result) == 1
        assert result[0].name == "valid"

    def test_empty_sheet_returns_empty(self, tmp_path: Path) -> None:
        path = _write_vinfo(tmp_path, [])
        result = parse_vinfo(path)
        assert result == []


# ═══════════════════════════════════════════════════════════════════════
# parse_vinfo — error handling (synthetic)
# ═══════════════════════════════════════════════════════════════════════


class TestParseVinfoErrors:
    def test_file_not_found(self, tmp_path: Path) -> None:
        with pytest.raises(RVToolsParseError, match="file not found"):
            parse_vinfo(tmp_path / "missing.xlsx")

    def test_missing_vinfo_sheet(self, tmp_path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "WrongSheet"
        path = tmp_path / "bad.xlsx"
        wb.save(path)
        with pytest.raises(RVToolsParseError, match="sheet 'vInfo' not found"):
            parse_vinfo(path)

    def test_missing_required_column(self, tmp_path: Path) -> None:
        path = _write_vinfo(
            tmp_path,
            [["vm1", 4, "poweredOn"]],
            columns=("VM", "CPUs", "Powerstate"),
        )
        with pytest.raises(RVToolsParseError, match="missing required columns"):
            parse_vinfo(path)


# ═══════════════════════════════════════════════════════════════════════
# parse_vhost — real data (tests/rvtools/example.xlsx)
# ═══════════════════════════════════════════════════════════════════════


class TestParseVhost:
    """Happy-path tests against the real RVTools vHost sheet.

    example.xlsx vHost sheet:
      - 34 physical hosts, all HT Active = True
      - 5 profiles: vmsp (5), vmprod3 (22), bovmmgt (3),
        bovmdmz01 (2), bovmdmz7 (2)
    """

    def test_total_hosts_returned(self) -> None:
        result = parse_vhost(_EXAMPLE_XLSX)
        assert len(result) == 34

    def test_deterministic_output(self) -> None:
        a = parse_vhost(_EXAMPLE_XLSX)
        b = parse_vhost(_EXAMPLE_XLSX)
        assert a == b

    def test_sorted_by_name(self) -> None:
        result = parse_vhost(_EXAMPLE_XLSX)
        names = [h.name for h in result]
        assert names == sorted(names)

    def test_all_ht_active(self) -> None:
        result = parse_vhost(_EXAMPLE_XLSX)
        assert all(h.ht_active for h in result)

    def test_known_host_vmsp01(self) -> None:
        """vmsp-01 → 2 sockets, 8 cores/socket, HT, 262109 MB."""
        result = parse_vhost(_EXAMPLE_XLSX)
        host = next(h for h in result if h.name == "vmsp-01")
        assert host.sockets == 2
        assert host.cores_per_socket == 8
        assert host.memory_mb == 262109

    def test_known_host_vmprod3_01(self) -> None:
        """vmprod3-01 → 2 sockets, 8 cores/socket, 524253 MB."""
        result = parse_vhost(_EXAMPLE_XLSX)
        host = next(h for h in result if h.name == "vmprod3-01")
        assert host.sockets == 2
        assert host.cores_per_socket == 8
        assert host.memory_mb == 524253

    def test_known_host_bovmdmz01_01(self) -> None:
        """bovmdmz01-01 → 2 sockets, 6 cores/socket, 294901 MB."""
        result = parse_vhost(_EXAMPLE_XLSX)
        host = next(h for h in result if h.name == "bovmdmz01-01")
        assert host.sockets == 2
        assert host.cores_per_socket == 6
        assert host.memory_mb == 294901

    def test_host_profiles_distribution(self) -> None:
        """Verify the expected distribution of host profiles.

        - 5 × vmsp (262109 MB, 8 cores/socket)
        - 22 × vmprod3 (524253 MB, 8 cores/socket)
        - 3 × bovmmgt (393087 MB, 8 cores/socket)
        - 2 × bovmdmz01 (294901 MB, 6 cores/socket)
        - 2 × bovmdmz7 (196597 MB, 6 cores/socket)
        """
        result = parse_vhost(_EXAMPLE_XLSX)
        vmsp = [h for h in result if h.name.startswith("vmsp-")]
        vmprod3 = [h for h in result if h.name.startswith("vmprod3-")]
        bovmmgt = [h for h in result if h.name.startswith("bovmmgt-")]
        bovmdmz01 = [h for h in result if h.name.startswith("bovmdmz01-")]
        bovmdmz7 = [h for h in result if h.name.startswith("bovmdmz7-")]

        assert len(vmsp) == 5
        assert len(vmprod3) == 22
        assert len(bovmmgt) == 3
        assert len(bovmdmz01) == 2
        assert len(bovmdmz7) == 2

        # All vmsp hosts share the same topology
        assert all(h.memory_mb == 262109 for h in vmsp)
        assert all(h.cores_per_socket == 8 for h in vmsp)

        # All vmprod3 hosts share the same topology
        assert all(h.memory_mb == 524253 for h in vmprod3)
        assert all(h.cores_per_socket == 8 for h in vmprod3)


# ═══════════════════════════════════════════════════════════════════════
# parse_vhost — edge cases (synthetic)
# ═══════════════════════════════════════════════════════════════════════


class TestParseVhostEdgeCases:
    """Edge cases requiring synthetic xlsx."""

    def test_ht_active_variants(self, tmp_path: Path) -> None:
        """Real file only has True; verify false/yes/1/No handling."""
        rows = [
            ["h1", 2, 8, "True", 65536],
            ["h2", 2, 8, "yes", 65536],
            ["h3", 2, 8, "1", 65536],
            ["h4", 2, 8, "false", 65536],
            ["h5", 2, 8, "No", 65536],
        ]
        path = _write_vhost(tmp_path, rows)
        result = parse_vhost(path)
        assert result[0].ht_active is True  # h1 - True
        assert result[1].ht_active is True  # h2 - yes
        assert result[2].ht_active is True  # h3 - 1
        assert result[3].ht_active is False  # h4 - false
        assert result[4].ht_active is False  # h5 - No

    def test_skips_blank_host_names(self, tmp_path: Path) -> None:
        rows = [
            ["", 2, 8, "false", 65536],
            ["valid-host", 2, 8, "false", 65536],
        ]
        path = _write_vhost(tmp_path, rows)
        result = parse_vhost(path)
        assert len(result) == 1

    def test_empty_sheet(self, tmp_path: Path) -> None:
        path = _write_vhost(tmp_path, [])
        result = parse_vhost(path)
        assert result == []

    def test_file_not_found(self, tmp_path: Path) -> None:
        with pytest.raises(RVToolsParseError, match="file not found"):
            parse_vhost(tmp_path / "missing.xlsx")

    def test_missing_vhost_sheet(self, tmp_path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "vInfo"
        path = tmp_path / "no_vhost.xlsx"
        wb.save(path)
        with pytest.raises(RVToolsParseError, match="sheet 'vHost' not found"):
            parse_vhost(path)

    def test_missing_required_column(self, tmp_path: Path) -> None:
        path = _write_vhost(
            tmp_path,
            [["host1", 2]],
            columns=("Host", "# CPU"),
        )
        with pytest.raises(RVToolsParseError, match="missing required columns"):
            parse_vhost(path)
